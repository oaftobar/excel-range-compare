import re
import openpyxl
from openpyxl.styles import Font, numbers

wb = openpyxl.load_workbook('Employee Sales.xlsx')

while True:
    print(f"Available worksheets: {wb.sheetnames}")
    ws1_input = input("Enter the name of Worksheet 1: ")
    while ws1_input not in wb.sheetnames:
        print("Worksheet does not exist. Please try again")
        ws1_input = input("Enter the name of Worksheet 1: ")
    ws2_input = input("Enter the name of Worksheet 2: ")
    while ws2_input not in wb.sheetnames:
        print("Worksheet does not exist. Please try again")
        ws2_input = input("Enter the name of Worksheet 2: ")

    print(f"Available Cell Ranges: {wb[ws1_input].dimensions}")
    rng_pattern = r'^[A-Z]{1,3}\d+:[A-Z]{1,3}\d+$'
    rng1_input = input("Enter cell range: ").upper()
    while re.search(rng_pattern, rng1_input) is None:
        print("Invalid Cell pattern.")
        rng1_input = input("Enter cell range: ").upper()
    rng2_input = input("Enter matching cell range: ").upper()
    while rng2_input != rng1_input:
        print("Cell range did not match.")
        rng2_input = input("Enter matching cell range: ").upper()

    ws1 = wb[ws1_input]
    ws2 = wb[ws2_input]

    rng1 = ws1[rng1_input]
    rng2 = ws2[rng2_input]

    num_rows = len(rng1)
    num_cols = len(rng1[0])
    
    if len(rng1) == len(rng2) and len(rng1[0]) == len(rng2[0]):
        break
    else:
        print("ERROR: Dimensions do not match")
        continue

headers = rng1[0]

diffs = []

for i in range(1, num_rows):
    for j in range(num_cols):
        cell1 = rng1[i][j].value
        cell2 = rng2[i][j].value

        if cell1 != cell2:
            diff = [i, headers[j].value, cell1, cell2]

            diffs.append(diff)

if 'Diffs' in wb.sheetnames:
    del wb['Diffs']
diffs_ws = wb.create_sheet('Diffs')

diffs_rng = diffs_ws['A2:D' + str(len(diffs) + 1)]

for i, row in enumerate(diffs_rng):
    for j, cell in enumerate(row):
        cell.value = diffs[i][j]

diffs_header_row = diffs_ws['A1:E1']

diffs_header = ['Row', 'Column', 'Value 1', 'Value 2', 'Delta']

for idx, cell in enumerate(diffs_header_row[0]):
    cell.value = diffs_header[idx]

delta_rng = diffs_ws['E2:E' + str(len(diffs) + 1)]

for idx, row in enumerate(delta_rng):
    current_row = str(idx + 2)

    diff_data = diffs[idx]

    if type(diff_data[2]) in(int, float) and type(diff_data[3]) in(int, float):
        cell1_addr = 'C' + current_row
        cell2_addr = 'D' + current_row
        formula = f"=ABS({cell1_addr}-{cell2_addr})/AVERAGE({cell1_addr}:{cell2_addr})"

        target_cell = row[0]
        target_cell.value = formula
        target_cell.font = Font(italic=True, color='FF0000')
        target_cell.number_format = numbers.FORMAT_PERCENTAGE

wb.save('Employee Sales NEW.xlsx')